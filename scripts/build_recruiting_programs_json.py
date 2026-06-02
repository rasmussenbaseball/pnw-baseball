#!/usr/bin/env python3
"""
Build backend/app/data/recruiting_programs.json from the hand-researched
"PNW Recruiting Guide.xlsx" so it can seed the recruiting_programs table.

Merges the per-division sheets (D1/D2/D3/NAIA/NWAC — facilities/stadium detail)
with the "All Schools" sheet (coach bio + email) by school name, normalizes every
column to a camelCase `profile` key, drops empty / "N/A" values (so the page can
hide them), and resolves each school to its NWBB team_id via the existing
RECRUIT_SCHOOLS map in frontend/src/data/recruitQuiz.js.

Usage:
    python3 scripts/build_recruiting_programs_json.py [path-to-xlsx]
Default xlsx: ~/Downloads/PNW Recruiting Guide.xlsx
"""
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "PNW Recruiting Guide.xlsx"
RECRUIT_QUIZ = REPO / "frontend" / "src" / "data" / "recruitQuiz.js"
OUT = REPO / "backend" / "app" / "data" / "recruiting_programs.json"

# Normalized header (lower, collapsed spaces) -> camelCase profile key.
HEADER_MAP = {
    "division": "division", "conference": "conference",
    "school name": "schoolName", "team name": "teamName",
    "city": "city", "state": "state",
    "head coach": "coach", "assistant coaches": "assistants",
    "coach bio": "coachBio", "coaching staff size": "staffSize", "coach email": "coachEmail",
    "alma mater (hc)": "coachAlma", "years at school (hc)": "coachYears",
    "previous coaching stops (hc)": "prevStops", "previous coaching stop": "prevStops",
    "career record at school": "careerRecord",
    "school type": "schoolType", "undergraduate enrollment": "enrollment",
    "student-to-faculty ratio": "sfr", "acceptance rate": "acceptance",
    "top 5 majors": "topMajors", "all majors url": "allMajorsUrl",
    "in-state tuition": "inStateTuition", "out-of-state tuition": "outStateTuition",
    "room & board": "roomBoard", "financial aid %": "financialAidPct",
    "average class size": "avgClassSize",
    "2026 roster size": "rosterSize", "2026 graduating seniors": "gradSeniors",
    "recent record": "recentRecord", "scholarship info": "scholarshipInfo",
    "recruiting questionnaire url": "recruitQuestionnaireUrl",
    "recruiting questionnaire": "recruitQuestionnaireUrl",
    "camp/clinic info": "campClinicInfo", "prospect day info": "prospectDayInfo",
    "stadium/field name": "stadium", "stadium capacity": "capacity",
    "field surface": "fieldSurface", "indoor facility": "indoorFacility",
    "campus setting": "campusSetting", "nearest airport": "nearestAirport",
    "distance from major city": "distanceFromCity", "distance from major ci": "distanceFromCity",
    "campus safety rating": "campusSafety", "graduation rate": "gradRate",
    "athletic academic support": "athleticAcademicSupport",
    "athletics website": "athleticsWebsite", "baseball roster url": "baseballRosterUrl",
    "school website": "schoolWebsite", "research status": "researchStatus", "notes": "notes",
}
DIVISION_SHORT = {"ncaa d1": "D1", "ncaa d2": "D2", "ncaa d3": "D3", "naia": "NAIA", "nwac": "NWAC"}
EMPTY = {"", "n/a", "none", "tbd", "-", "--", "---"}


def norm_header(h):
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def key_for(header):
    nh = norm_header(header)
    if nh in HEADER_MAP:
        return HEADER_MAP[nh]
    for label, key in HEADER_MAP.items():  # prefix fallback (truncated headers)
        if nh.startswith(label) or label.startswith(nh):
            return key
    return None


def clean(v):
    if v is None:
        return None
    # openpyxl returns numeric cells as floats; show 20622, not "20622.0".
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = re.sub(r"\s+", " ", str(v).strip())
    if s.lower() in EMPTY:
        return None
    return s


def load_team_ids():
    """name -> teamId from RECRUIT_SCHOOLS in recruitQuiz.js."""
    text = RECRUIT_QUIZ.read_text()
    start = text.index("RECRUIT_SCHOOLS")
    body = text[start:]
    pairs = re.findall(r'"name":\s*"([^"]+)"[\s\S]*?"teamId":\s*(\d+)', body)
    return {name.strip().lower(): int(tid) for name, tid in pairs}


def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {}
        for i, v in enumerate(r):
            if i >= len(header):
                continue
            k = key_for(header[i])
            cv = clean(v)
            if k == "rosterSize" and cv == "0":
                cv = None  # 0 is a placeholder for "unknown", not a real roster size
            if k and cv is not None:
                rec[k] = cv
        out.append(rec)
    return out


def main():
    if not XLSX.exists():
        print(f"ERROR: xlsx not found at {XLSX}", file=sys.stderr)
        sys.exit(1)
    team_ids = load_team_ids()
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    # Per-division sheets are canonical (most complete); All Schools fills bio/email gaps.
    per_div = []
    for s in ("D1", "D2", "D3", "NAIA", "NWAC"):
        if s in wb.sheetnames:
            per_div += sheet_records(wb[s])
    all_schools = {}
    if "All Schools" in wb.sheetnames:
        for rec in sheet_records(wb["All Schools"]):
            nm = (rec.get("schoolName") or "").strip().lower()
            if nm:
                all_schools[nm] = rec

    programs, unmatched = [], []
    for rec in per_div:
        school = (rec.get("schoolName") or "").strip()
        if not school:
            continue
        # Overlay any field present in All Schools but missing here (notably coachBio/coachEmail).
        overlay = all_schools.get(school.lower(), {})
        for k, v in overlay.items():
            if k not in rec or rec[k] in (None, ""):
                rec[k] = v
        tid = team_ids.get(school.lower())
        if tid is None:
            unmatched.append(school)
            continue
        division = DIVISION_SHORT.get(norm_header(rec.get("division")), rec.get("division"))
        rec["division"] = division
        programs.append({
            "team_id": tid,
            "school_name": school,
            "division": division,
            "conference": rec.get("conference"),
            "profile": rec,
        })

    programs.sort(key=lambda p: (p["division"] or "", p["school_name"]))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(programs, indent=2, ensure_ascii=False))

    print(f"Wrote {len(programs)} programs -> {OUT}")
    if unmatched:
        print(f"UNMATCHED (no team_id): {unmatched}", file=sys.stderr)
    # Coverage by division
    from collections import Counter
    print("By division:", dict(Counter(p["division"] for p in programs)))


if __name__ == "__main__":
    main()
